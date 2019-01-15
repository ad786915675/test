#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.drawing.image import Image
from PIL import Image
import os

inexcelname='1.xlsx'
outexcelname='111.xlsx'

def run():
    path=os.path.abspath('.')
    picpath=os.path.join(path,'pic')  #图片路径
    docpath=os.path.join(path,'doc')

    infile=os.path.join(docpath,inexcelname)
    if(not os.path.exists(infile)):
        print("原文件不存在："+infile)
        return
    wb = load_workbook(infile)
    ws = wb.active
    rows = ws.rows

    l=[]
    for row in rows:
        line=[col.value for col in row]
        l.append(line)

    outfile=os.path.join(path,outexcelname)
    if(os.path.exists(outfile)):
        print("Excel文件已经存在："+outfile)
        return
    book = xlsxwriter.Workbook(outfile)
    sheet = book.add_worksheet('sheet1')

    row=0
    for root,dirs,files in os.walk(picpath):
        for line in l:
            column=0
            if(row==0 or row==1):
                for c in line:
                    sheet.write(row,column,c)
                    column+=1
            else:
                for c in line:
                    if(column==2):
                        name=c
                    if(column==3):
                        filename=name+'.jpg'
                        if(filename in files):
                            filename=os.path.join(root,filename)
                            sheet.insert_image(row,column,filename,{'x_scale': 0.125, 'y_scale': 0.125})
                        else:
                            print(name+" 没有照片！")
                    else:
                        sheet.write(row,column,c)
                    column+=1
            row+=1

    book.close()
    print("已生成Excel文件："+outfile)

if __name__ == '__main__':
    run()

