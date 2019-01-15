#coding=utf-8

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

wb = load_workbook("./jing.xlsx")
ws = wb.active
rows = ws.rows

l=[]
for row in rows:
    line=[col.value for col in row][:14]
    l.append(line)
l=l[:4273]
names=[]
sheet=l
for line in sheet:
    names.append(line[3])

names=set(names)

result=[]
ct=0
for name in names:
    temp=[]
    tmp=[]
    for line in sheet:
        if(line[3]==name):
            temp.append(line)
    flag=0
    for t in temp:
        if(t[4]=='人民币'):
            flag+=1
    if(flag>0):
        tmp=[]
        f=0
        value=[]
        for t in temp:
            if(t[4]=='人民币'):
                tmp.append(t)
                value.append(t[5])
                f+=1
        if(f>0):
            tall=max(value)
            tt=[]
            for t in tmp:
                if(t[4]=='人民币' and t[5]==tall):
                    tt.append(t)
            tmp=tt
    else:
        tmp=temp
    print(name+str(len(tmp)))
    ct+=len(tmp)
    for t in tmp:
        result.append(t)

# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb2 = Workbook()
#获取当前活跃的worksheet,默认就是第一个worksheet
ws2 = wb2.active

for l in result:
    ws2.append(l)

wb2.save(filename="./2.xlsx")













